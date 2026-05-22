import io
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from difflib import SequenceMatcher
import pandas as pd
import streamlit as st

TZ = ZoneInfo("America/New_York")
CACHE_TTL_SECONDS = 2 * 60 * 60

st.set_page_config(page_title="BT Session Start Checker", layout="wide")
st.title("BT Session Start Checker")

with st.expander("Instructions"):
    st.markdown("""
    This tool checks whether **BT session start times in HiRasmus** match the **scheduled start times in Aloha** within a chosen tolerance window.

    ---

    ### 🔹 Files to Upload

    1. **File 1 – HiRasmus export**  
    A CSV or Excel file that contains one row per session attempt. It **must include** at least these columns (case-sensitive):

    - `Start time` – the actual time the BT clicked "Start" in HiRasmus  
    - `Aloha Appointment ID` – the appointment ID that links to Aloha  
    - `Status` – current session status (e.g., Completed, In progress, etc.)

    2. **File 2 – Aloha Sessions export**  
    A CSV or Excel file with the scheduled sessions from Aloha. It **must include** at least:

    - `Staff Name` – BT’s name as it appears in Aloha  
    - `Client Name` – client’s name  
    - `Appointment ID` – appointment ID (will be matched to `Aloha Appointment ID`)  
    - `Appt. Start Time` – scheduled start time (e.g., 3:00 PM)  
    - `Service Name` – service type (the tool only checks rows where this is `Direct Service BT`)  
    - `Client City` – used for filtering or reporting

    3. **(Optional) BT Contacts file**
    You may also upload a separate **BT contact list** in CSV/Excel format with columns:

    - `BT Name` – BT’s name (e.g., "Jane Doe")
    - `Phone` – BT’s phone number
    - `Email` – BT’s email

    The app will try to **fuzzy-match** `BT Name` to the Aloha `Staff Name` and attach Phone/Email to each row, so you can easily follow up with BTs on flagged sessions.

    4. **(Optional) Case Coordinator file**
    You may also upload a separate **Case Coordinator list** in CSV/Excel format with columns:

    - `Child's Full Name` – child's name in "First Last" format (e.g., "Jane Doe")
    - `Case Coordinator Name` – coordinator assigned to the child
    - `Guardian 1 Phone` – guardian's phone number (optional)

    The app will try to **fuzzy-match** `Child's Full Name` to the Aloha `Client Name` and attach the `Case Coordinator Name` and `Guardian 1 Phone` (labelled `Guardian Number` in the output) to each row.

    ---

    ### 🔍 What this checker does

    - Converts and aligns **scheduled start time** from Aloha and **actual start time** from HiRasmus to the same timezone (America/New_York).
    - For each `Aloha Appointment ID`, it keeps the **earliest start** recorded in HiRasmus.
    - Calculates the time difference (in minutes) between **scheduled** and **actual** start times.
    - Classifies each session as:
      - **Within tolerance** – actual start is inside your ± *Tolerance (minutes)* setting.
      - **Outside tolerance** – actual start is more than that tolerance early/late.
      - **No start recorded** – no HiRasmus start time found.
      - **Future session** – scheduled start time is in the future.
      - **Missing/invalid time** – Appt. Start Time is missing or in an unrecognized format.

    The results are split into three tables and Excel sheets:
    - **Valid** – within tolerance  
    - **Flagged** – need review (late/early/missing starts, etc.)  
    - **Future** – sessions that haven’t started yet

    Use the options in the left sidebar to adjust the tolerance window, restrict to the **current minute only**, and choose whether to **assume today’s date** when the Aloha file has no appointment date.
    """)


with st.sidebar:
    st.header("Options")
    tolerance_min = st.number_input("Tolerance (± minutes)", min_value=0, max_value=120, value=5, step=1)
    only_check_current_minute = st.checkbox("Only check the current minute", value=False)
    assume_today_if_no_date = st.checkbox("If no Appt. Date, assume today", value=True)
    show_tables = st.checkbox("Preview results in app", value=True)

    st.markdown("---")
    st.header("BT Contacts (optional)")
    bt_contacts_file = st.file_uploader(
        "Upload BT Contacts (BT Name, Phone, Email)",
        type=["csv", "xlsx"],
        key="bt_contacts"
    )

    st.markdown("---")
    st.header("Case Coordinators (optional)")
    coordinators_file = st.file_uploader(
        "Upload Case Coordinators (Child's Full Name, Case Coordinator Name)",
        type=["csv", "xlsx"],
        key="coordinators"
    )

col1, col2 = st.columns(2)
with col1:
    f_hirasmus = st.file_uploader("Upload File 1: HiRasmus", type=["csv", "xlsx"])
with col2:
    f_aloha = st.file_uploader("Upload File 2: Aloha Sessions", type=["csv", "xlsx"])


def read_any(file):
    if file is None:
        return None
    if file.name.lower().endswith(".csv"):
        return pd.read_csv(file, dtype=str)
    return pd.read_excel(file, dtype=str, engine="openpyxl")


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].astype(str).str.strip()
    return df


def to_last_first(name: str) -> str:
    # Convert "First Last" -> "Last, First"
    parts = str(name).strip().split()
    if len(parts) >= 2:
        first = " ".join(parts[:-1])
        last = parts[-1]
        return f"{last}, {first}"
    return str(name).strip()


def norm_name(s: str) -> str:
    # Normalize strings for fuzzy comparison
    s = str(s).strip().lower()
    s = s.replace(",", " ")
    s = " ".join(s.split())  # collapse multiple spaces
    return s


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def match_bt_contacts(staff_names, bt_candidates, threshold=0.8):
    matches = {}
    for staff in staff_names:
        staff_norm = norm_name(staff)
        best_score = 0.0
        best_contact = None
        for bt_name_norm, phone, email in bt_candidates:
            score = SequenceMatcher(None, staff_norm, bt_name_norm).ratio()
            if score > best_score:
                best_score = score
                best_contact = (phone, email)
        if best_contact is not None and best_score >= threshold:
            matches[staff] = best_contact
    return matches


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def match_case_coordinators(client_names, coord_candidates, threshold=0.8):
    matches = {}
    for client in client_names:
        client_norm = norm_name(client)
        best_score = 0.0
        best_coord = None
        for child_name_norm, coordinator, guardian in coord_candidates:
            score = SequenceMatcher(None, client_norm, child_name_norm).ratio()
            if score > best_score:
                best_score = score
                best_coord = (coordinator, guardian)
        if best_coord is not None and best_score >= threshold:
            matches[client] = best_coord
    return matches


def parse_time_cell(val):
    s = (str(val) if val is not None else "").strip()
    for fmt in ("%I:%M:%S %p", "%I:%M %p", "%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except Exception:
            pass
    return None


def round_to_minute(dt_obj):
    if pd.isna(dt_obj):
        return dt_obj
    return dt_obj.replace(second=0, microsecond=0)


def drop_tz(series):
    s = pd.to_datetime(series, errors="coerce")
    try:
        return s.dt.tz_convert(TZ).dt.tz_localize(None)
    except Exception:
        return s.dt.tz_localize(None) if getattr(s.dt, "tz", None) is not None else s


def local_now():
    return datetime.now(TZ).replace(second=0, microsecond=0)


# ---------- MAIN LOGIC ----------
if f_hirasmus is not None and f_aloha is not None:
    try:
        df_hi = read_any(f_hirasmus)
        df_al = read_any(f_aloha)

        df_hi = normalize_cols(df_hi)
        df_al = normalize_cols(df_al)

        # Validate required columns
        req_hi = {"Start time", "AlohaABA Appointment ID", "Status"}
        req_al = {"Staff Name", "Client Name", "Appointment ID", "Appt. Start Time", "Service Name", "Client City"}

        missing_hi = req_hi - set(df_hi.columns)
        missing_al = req_al - set(df_al.columns)
        if missing_hi:
            st.error(f"HiRasmus is missing: {sorted(missing_hi)}")
            st.stop()
        if missing_al:
            st.error(f"Aloha Sessions is missing: {sorted(missing_al)}")
            st.stop()

        # Fix ID types (remove .0 from floats / Excel)
        df_hi["Aloha Appointment ID"] = (
            df_hi["AlohaABA Appointment ID"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        )
        df_al["Appointment ID"] = df_al["Appointment ID"].astype(str).str.strip()

        # Filter Aloha to "Direct Service BT"
        df_al = df_al[df_al["Service Name"] == "Direct Service BT"].copy()

        # Exclude demo/test clients
        EXCLUDED_CLIENTS = {"Wang Demo, Marry"}
        df_al = df_al[~df_al["Client Name"].isin(EXCLUDED_CLIENTS)].copy()

        df_al.reset_index(drop=True, inplace=True)

        # Discover / normalize the Appt. End Time column (if present)
        end_time_cols = [
            c for c in df_al.columns
            if c.strip().lower() in {"appt. end time", "appointment end time", "end time"}
        ]
        if end_time_cols:
            if end_time_cols[0] != "Appt. End Time":
                df_al = df_al.rename(columns={end_time_cols[0]: "Appt. End Time"})
        else:
            df_al["Appt. End Time"] = ""

        # -----------------------------
        # OPTIONAL: Fuzzy match BT contacts -> Staff Name
        # -----------------------------
        df_al["BT Number"] = ""
        df_al["Email"] = ""

        if bt_contacts_file is not None:
            bt_df = read_any(bt_contacts_file)
            bt_df = normalize_cols(bt_df)

            bt_required = {"BT Name", "Phone", "Email"}
            bt_missing = bt_required - set(bt_df.columns)
            if bt_missing:
                st.error(f"BT Contacts file is missing: {sorted(bt_missing)}")
            else:
                bt_df["BT_formatted"] = bt_df["BT Name"].apply(to_last_first)
                bt_df["bt_norm"] = bt_df["BT_formatted"].apply(norm_name)

                staff_unique = tuple(df_al["Staff Name"].dropna().unique())
                bt_candidates = tuple(
                    bt_df[["bt_norm", "Phone", "Email"]].itertuples(index=False, name=None)
                )
                staff_matches = match_bt_contacts(staff_unique, bt_candidates)
                staff_to_phone = {staff: phone for staff, (phone, _) in staff_matches.items()}
                staff_to_email = {staff: email for staff, (_, email) in staff_matches.items()}

                # Attach to df_al (Phone -> BT Number)
                df_al["BT Number"] = df_al["Staff Name"].map(staff_to_phone)
                df_al["Email"] = df_al["Staff Name"].map(staff_to_email)

        # -----------------------------
        # OPTIONAL: Fuzzy match Case Coordinators -> Client Name
        # -----------------------------
        df_al["Case Coordinator Name"] = ""
        df_al["Guardian Number"] = ""

        if coordinators_file is not None:
            coord_df = read_any(coordinators_file)
            coord_df = normalize_cols(coord_df)

            coord_required = {"Child's Full Name", "Case Coordinator Name"}
            coord_missing = coord_required - set(coord_df.columns)
            if coord_missing:
                st.error(f"Case Coordinator file is missing: {sorted(coord_missing)}")
            else:
                has_guardian = "Guardian 1 Phone" in coord_df.columns

                # Child's Full Name is "First Last" -> convert to "Last, First"
                # to line up with Aloha Client Name formatting
                coord_df["Child_formatted"] = coord_df["Child's Full Name"].apply(to_last_first)
                coord_df["child_norm"] = coord_df["Child_formatted"].apply(norm_name)

                if not has_guardian:
                    coord_df["Guardian 1 Phone"] = ""

                client_unique = tuple(df_al["Client Name"].dropna().unique())
                coord_candidates = tuple(
                    coord_df[
                        ["child_norm", "Case Coordinator Name", "Guardian 1 Phone"]
                    ].itertuples(index=False, name=None)
                )
                client_matches = match_case_coordinators(client_unique, coord_candidates)
                client_to_coord = {
                    client: coordinator
                    for client, (coordinator, _) in client_matches.items()
                }
                client_to_guardian = {
                    client: guardian
                    for client, (_, guardian) in client_matches.items()
                }

                # Attach to df_al
                df_al["Case Coordinator Name"] = df_al["Client Name"].map(client_to_coord)
                if has_guardian:
                    df_al["Guardian Number"] = df_al["Client Name"].map(client_to_guardian)

        # Try to get Appt. Date
        date_cols = [c for c in df_al.columns if c.strip().lower() in {"appt. date", "appointment date", "date"}]
        appt_date_col = date_cols[0] if date_cols else None

        if appt_date_col:
            appt_dates = pd.to_datetime(df_al[appt_date_col], errors="coerce")
        else:
            appt_dates = pd.Series(pd.NaT, index=df_al.index, dtype="datetime64[ns]")

        today_local = local_now().date()
        scheduled_list = []
        for pos in range(len(df_al)):
            row = df_al.iloc[pos]
            t = parse_time_cell(row.get("Appt. Start Time", None))
            if t is None:
                scheduled_list.append(pd.NaT)
                continue
            d = appt_dates.iloc[pos]
            if pd.isna(d):
                if assume_today_if_no_date:
                    d = today_local
                else:
                    scheduled_list.append(pd.NaT)
                    continue
            else:
                d = d.date()
            dt_local = datetime.combine(d, t).replace(tzinfo=TZ)
            scheduled_list.append(dt_local)

        df_al["_scheduled_dt"] = pd.Series([round_to_minute(dt) for dt in scheduled_list], index=df_al.index)

        # Localize Start Time - strip microseconds first
        df_hi["Start Time Clean"] = df_hi["Start time"].astype(str).str.replace(r"\.[\d]+$", "", regex=True)
        start_times = pd.to_datetime(df_hi["Start Time Clean"], errors="coerce")
        processed_times = []
        for dt in start_times:
            if pd.isna(dt):
                processed_times.append(pd.NaT)
            else:
                dt_obj = dt.to_pydatetime() if hasattr(dt, "to_pydatetime") else dt
                if dt_obj.tzinfo is not None:
                    dt_obj = dt_obj.astimezone(TZ)
                else:
                    dt_obj = dt_obj.replace(tzinfo=TZ)
                processed_times.append(round_to_minute(dt_obj))
        df_hi["_actual_start_dt"] = pd.Series(processed_times, index=df_hi.index)

        # Keep earliest per Aloha Appointment ID
        df_hi_sorted = df_hi.sort_values(by=["Aloha Appointment ID", "_actual_start_dt"])
        df_hi_min = df_hi_sorted.groupby("Aloha Appointment ID", as_index=False).first()

        # Merge on Appointment ID (Aloha) vs Aloha Appointment ID (HiRasmus)
        merged = pd.merge(
            df_al,
            df_hi_min[["Aloha Appointment ID", "_actual_start_dt", "Start time", "Status"]],
            left_on="Appointment ID",
            right_on="Aloha Appointment ID",
            how="left",
        )

        # Normalize timezones again — force both columns to be tz-aware in TZ.
        # This handles the all-NaT case (where pd.to_datetime would otherwise
        # return naive datetime64 and break tz-aware subtraction).
        def _force_tz_aware(series, tz):
            s = pd.to_datetime(series, errors="coerce")
            # Object dtype (mixed tz-aware Timestamps) — normalize element-wise
            if s.dtype == object:
                def fix(x):
                    if pd.isna(x):
                        return pd.NaT
                    if getattr(x, "tzinfo", None) is None:
                        return x.tz_localize(tz) if hasattr(x, "tz_localize") else x.replace(tzinfo=tz)
                    return x.tz_convert(tz) if hasattr(x, "tz_convert") else x.astimezone(tz)
                return s.apply(fix)
            # datetime64 dtype path
            try:
                if s.dt.tz is None:
                    return s.dt.tz_localize(tz, nonexistent="NaT", ambiguous="NaT")
                return s.dt.tz_convert(tz)
            except (AttributeError, TypeError):
                return s

        merged["_scheduled_dt"] = _force_tz_aware(merged["_scheduled_dt"], TZ)
        merged["_actual_start_dt"] = _force_tz_aware(merged["_actual_start_dt"], TZ)

        # Compute minutes_diff defensively (works for both datetime64 and object dtype).
        def _diff_minutes(actual, scheduled):
            if pd.isna(actual) or pd.isna(scheduled):
                return float("nan")
            return (actual - scheduled).total_seconds() / 60.0

        merged["minutes_diff"] = [
            _diff_minutes(a, s)
            for a, s in zip(merged["_actual_start_dt"], merged["_scheduled_dt"])
        ]

        now_local = local_now()

        to_check = merged.copy()
        if only_check_current_minute:
            to_check = to_check[to_check["_scheduled_dt"] == now_local]

        def reason_row(row):
            if pd.isna(row.get("_scheduled_dt")):
                raw_time = str(row.get("Appt. Start Time", "")).strip()
                return "Missing/invalid Appt. Start Time" if not raw_time else "Invalid Appt. Start Time format"
            if row["_scheduled_dt"] > now_local:
                return "Future session"
            if pd.isna(row.get("_actual_start_dt")):
                return "No start recorded (not started)"
            if pd.isna(row.get("minutes_diff")):
                return "Missing time difference"
            if abs(row["minutes_diff"]) > tolerance_min:
                delta = int(round(row["minutes_diff"]))
                return f"Outside tolerance (Δ={delta} min)"
            return "Within tolerance"

        to_check["Reason"] = to_check.apply(reason_row, axis=1)

        df_future = to_check[to_check["_scheduled_dt"] > now_local].copy()
        df_past_now = to_check[to_check["_scheduled_dt"] <= now_local].copy()
        df_valid = df_past_now[df_past_now["Reason"] == "Within tolerance"].copy()
        df_flagged = df_past_now[df_past_now["Reason"] != "Within tolerance"].copy()

        cols = [
            "Staff Name",
            "BT Number",
            "Email",
            "Client Name",
            "Case Coordinator Name",
            "Guardian Number",
            "Appointment ID",
            "Aloha Appointment ID",
            "Client City",
            "Service Name",
            "Appt. Start Time",
            "Start time",
            "Status",
            "_scheduled_dt",
            "_actual_start_dt",
            "minutes_diff",
            "Reason",
        ]

        def safe_cols(df_):
            return [c for c in cols if c in df_.columns]

        out_valid = df_valid[safe_cols(df_valid)]
        out_flagged = df_flagged[safe_cols(df_flagged)]
        out_future = df_future[safe_cols(df_future)]

        if show_tables:
            st.subheader("✅ Valid Sessions")
            st.dataframe(out_valid, use_container_width=True, hide_index=True)
            st.subheader("⚠️ Flagged Sessions")
            st.dataframe(out_flagged, use_container_width=True, hide_index=True)
            st.subheader("📅 Future Sessions")
            st.dataframe(out_future, use_container_width=True, hide_index=True)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for name, df_export in {
                "Valid": out_valid,
                "Flagged": out_flagged,
                "Future": out_future,
            }.items():
                df_exp = df_export.copy()
                for col in ["_scheduled_dt", "_actual_start_dt"]:
                    if col in df_exp.columns:
                        df_exp[col] = drop_tz(df_exp[col])
                df_exp.to_excel(writer, index=False, sheet_name=name)
                ws = writer.sheets[name]
                for i, col in enumerate(df_exp.columns, 1):
                    max_len = min(48, max(df_exp[col].astype(str).str.len().max(), len(col)) + 2)
                    ws.column_dimensions[get_column_letter(i)].width = max_len

        st.download_button(
            "📥 Download Excel (Valid / Flagged / Future)",
            data=buffer.getvalue(),
            file_name=f"bt_session_check_{datetime.now(TZ).strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # -----------------------------
        # Secondary export: "No matching record" only
        # (ignores tolerance — just past-due sessions with no HiRasmus start)
        # -----------------------------
        no_match_cols = [
            "Staff Name",
            "BT Number",
            "Email",
            "Client Name",
            "Case Coordinator Name",
            "Guardian Number",
            "Appt. Start Time",
            "Appt. End Time",
            "Completed",
            "Remark",
        ]

        no_match_df = df_past_now[df_past_now["_actual_start_dt"].isna()].copy()

        # Add fillable columns
        no_match_df["Completed"] = "No"
        no_match_df["Remark"] = ""

        # Sort by Case Coordinator Name (blanks last)
        no_match_df["_coord_sort_key"] = (
            no_match_df["Case Coordinator Name"].fillna("").astype(str).str.strip()
        )
        no_match_df["_coord_blank"] = no_match_df["_coord_sort_key"] == ""
        no_match_df = no_match_df.sort_values(
            by=["_coord_blank", "_coord_sort_key", "Client Name"],
            kind="stable",
        ).reset_index(drop=True)

        # Insert a blank separator row whenever the coordinator changes
        rows_out = []
        prev_coord = None
        for _, r in no_match_df.iterrows():
            cur_coord = r["_coord_sort_key"]
            if prev_coord is not None and cur_coord != prev_coord:
                rows_out.append({c: "" for c in no_match_cols})  # blank separator row
            rows_out.append({c: r.get(c, "") for c in no_match_cols})
            prev_coord = cur_coord

        no_match_out = pd.DataFrame(rows_out, columns=no_match_cols)

        buffer_no_match = io.BytesIO()
        with pd.ExcelWriter(buffer_no_match, engine="openpyxl") as writer:
            no_match_out.to_excel(writer, index=False, sheet_name="No Matching Record")
            ws = writer.sheets["No Matching Record"]

            # Column widths
            for i, col in enumerate(no_match_out.columns, 1):
                if len(no_match_out) == 0:
                    max_len = len(col) + 2
                else:
                    max_len = min(48, max(no_match_out[col].astype(str).str.len().max(), len(col)) + 2)
                ws.column_dimensions[get_column_letter(i)].width = max_len

            # Yes/No dropdown on the Completed column
            if "Completed" in no_match_out.columns and len(no_match_out) > 0:
                completed_idx = list(no_match_out.columns).index("Completed") + 1  # 1-based
                completed_letter = get_column_letter(completed_idx)
                dv = DataValidation(
                    type="list",
                    formula1='"Yes,No"',
                    allow_blank=True,
                    showDropDown=False,  # False => dropdown arrow IS shown (openpyxl quirk)
                )
                dv.add(f"{completed_letter}2:{completed_letter}{len(no_match_out) + 1}")
                ws.add_data_validation(dv)

        st.download_button(
            "📥 Download Excel (No Matching Record only)",
            data=buffer_no_match.getvalue(),
            file_name=f"bt_no_matching_record_{datetime.now(TZ).strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.caption(
            f"Now: **{now_local.strftime('%Y-%m-%d %I:%M %p')}** | "
            f"Tolerance: ±{tolerance_min} min | "
            f"{'Current minute only' if only_check_current_minute else 'All sessions'} | "
            f"{'Assuming today if no date' if assume_today_if_no_date else ''}"
        )

    except Exception as e:
        st.error(f"🚨 Error processing files: {e}")

else:
    st.info("⬆️ Please upload both files to proceed.")
